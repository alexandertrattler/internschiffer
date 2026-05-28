#!/usr/bin/env python3
from __future__ import annotations

import json
import re
import shutil
from datetime import datetime
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
from pathlib import Path
from urllib.parse import urlparse

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - reported through /api/health
    load_workbook = None


ROOT = Path(__file__).resolve().parents[1]
AGENCIES_JSON = ROOT / "agencies.json"
DATA_JS = ROOT / "data.js"
BACKUP_DIR = ROOT / "backups"

FIELD_ALIASES = {
    "id": "id",
    "name": "name",
    "firma": "name",
    "agentur": "name",
    "company": "name",
    "providedurl": "providedUrl",
    "provided_url": "providedUrl",
    "original": "providedUrl",
    "eingabe": "providedUrl",
    "domain": "domain",
    "website": "domain",
    "webseite": "domain",
    "url": "sourceUrl",
    "sourceurl": "sourceUrl",
    "source_url": "sourceUrl",
    "quelle": "sourceUrl",
    "location": "location",
    "ort": "location",
    "stadt": "location",
    "standort": "location",
    "profile": "profile",
    "profil": "profile",
    "beschreibung": "profile",
    "confidence": "confidence",
    "sicherheit": "confidence",
}

STRING_FIELDS = ["name", "providedUrl", "domain", "location", "profile", "confidence", "sourceUrl"]


def clean_key(value: object) -> str:
    key = str(value or "").strip().lower()
    key = re.sub(r"[^a-z0-9_äöüß]+", "", key)
    return key.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss")


def clean_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def normalize_domain(value: str) -> str:
    value = value.strip()
    if not value:
        return ""
    value = re.sub(r"^https?://", "", value, flags=re.I)
    value = value.split("/")[0]
    return value.strip()


def source_from_domain(domain: str) -> str:
    if not domain:
        return ""
    first = domain.split(";")[0].strip()
    if not first or first.lower() == "missing":
        return ""
    return f"https://{first}"


def parse_json_upload(raw: bytes) -> list[dict]:
    text = raw.decode("utf-8-sig").strip()
    if text.startswith("window.AGENCIES"):
        text = re.sub(r"^window\.AGENCIES\s*=\s*", "", text).rstrip(";").strip()
    payload = json.loads(text)
    if isinstance(payload, dict):
        payload = payload.get("agencies", payload.get("data", payload))
    if not isinstance(payload, list):
        raise ValueError("JSON muss ein Array oder ein Objekt mit agencies/data enthalten.")
    return payload


def parse_xlsx_upload(raw: bytes) -> list[dict]:
    if load_workbook is None:
        raise ValueError("openpyxl ist nicht installiert; XLSX kann nicht gelesen werden.")
    workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel-Datei ist leer.")
    headers = [FIELD_ALIASES.get(clean_key(header), clean_key(header)) for header in rows[0]]
    records = []
    for row in rows[1:]:
        record = {}
        for index, value in enumerate(row):
            if index >= len(headers):
                continue
            key = headers[index]
            if not key:
                continue
            record[key] = value
        if any(clean_text(value) for value in record.values()):
            records.append(record)
    return records


def normalize_agencies(records: list[dict]) -> list[dict]:
    agencies = []
    for index, record in enumerate(records, start=1):
        raw_id = clean_text(record.get("id", ""))
        agency = {
            "id": int(raw_id) if raw_id.isdigit() else index,
        }
        for field in STRING_FIELDS:
            agency[field] = clean_text(record.get(field, ""))

        agency["domain"] = normalize_domain(agency["domain"])
        if not agency["sourceUrl"]:
            agency["sourceUrl"] = agency["providedUrl"] if agency["providedUrl"].startswith(("http://", "https://")) else source_from_domain(agency["domain"])
        if not agency["confidence"]:
            agency["confidence"] = "medium"

        if agency["name"]:
            agencies.append(agency)

    if not agencies:
        raise ValueError("Keine gueltigen Eintraege gefunden. Mindestens eine Spalte name/Firma wird benoetigt.")

    seen_ids = set()
    for index, agency in enumerate(agencies, start=1):
        if agency["id"] in seen_ids:
            agency["id"] = index
        seen_ids.add(agency["id"])
    return agencies


def backup_current_files() -> None:
    BACKUP_DIR.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    for path in (AGENCIES_JSON, DATA_JS):
        if path.exists():
            shutil.copy2(path, BACKUP_DIR / f"{path.stem}-{stamp}{path.suffix}")


def write_agency_files(agencies: list[dict]) -> None:
    rendered = json.dumps(agencies, ensure_ascii=False, indent=2)
    backup_current_files()
    AGENCIES_JSON.write_text(f"{rendered}\n", encoding="utf-8")
    DATA_JS.write_text(f"window.AGENCIES = {rendered};\n", encoding="utf-8")


def parse_multipart_file(content_type: str, body: bytes) -> tuple[str, bytes]:
    match = re.search(r"boundary=(?P<boundary>[^;]+)", content_type)
    if not match:
        raise ValueError("Upload muss multipart/form-data sein.")
    boundary = ("--" + match.group("boundary").strip('"')).encode()
    for part in body.split(boundary):
        if b'Content-Disposition:' not in part or b'name="file"' not in part:
            continue
        header, _, payload = part.partition(b"\r\n\r\n")
        name_match = re.search(rb'filename="([^"]*)"', header)
        filename = name_match.group(1).decode("utf-8", "replace") if name_match else "upload"
        return filename, payload.rstrip(b"\r\n-")
    raise ValueError("Keine Datei im Upload gefunden.")


class Handler(BaseHTTPRequestHandler):
    def end_headers(self) -> None:
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        super().end_headers()

    def do_OPTIONS(self) -> None:
        self.send_response(HTTPStatus.NO_CONTENT)
        self.end_headers()

    def do_GET(self) -> None:
        if urlparse(self.path).path != "/api/health":
            self.send_json({"error": "Not found"}, HTTPStatus.NOT_FOUND)
            return
        self.send_json({
            "ok": True,
            "xlsx": load_workbook is not None,
            "projectRoot": str(ROOT),
        })

    def do_POST(self) -> None:
        if urlparse(self.path).path != "/api/upload":
            self.send_json({"error": "Not found"}, HTTPStatus.NOT_FOUND)
            return
        try:
            length = int(self.headers.get("Content-Length", "0"))
            filename, raw = parse_multipart_file(self.headers.get("Content-Type", ""), self.rfile.read(length))
            if filename.lower().endswith(".json"):
                records = parse_json_upload(raw)
            elif filename.lower().endswith(".xlsx"):
                records = parse_xlsx_upload(raw)
            else:
                raise ValueError("Bitte eine .json oder .xlsx Datei hochladen.")
            agencies = normalize_agencies(records)
            write_agency_files(agencies)
            self.send_json({"ok": True, "count": len(agencies), "agencies": agencies})
        except Exception as error:  # noqa: BLE001 - API should report user-facing import errors
            self.send_json({"ok": False, "error": str(error)}, HTTPStatus.BAD_REQUEST)

    def send_json(self, payload: dict, status: HTTPStatus = HTTPStatus.OK) -> None:
        raw = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(raw)))
        self.end_headers()
        self.wfile.write(raw)


def main() -> None:
    server = ThreadingHTTPServer(("127.0.0.1", 8787), Handler)
    print("internschiffer backend: http://127.0.0.1:8787")
    server.serve_forever()


if __name__ == "__main__":
    main()
